"""
importar_habilidades.py
-----------------------
Lê agendamentos_normalizados.xlsx e insere os dados no banco do Django.

USO:
    python importar_habilidades.py
    python importar_habilidades.py --dry-run   # apenas mostra o que seria salvo
    python importar_habilidades.py --reset      # apaga tudo e reimporta
"""

import os
import sys
import django
import argparse
from datetime import datetime

# ------------------------------------------------------------------
# Configuração do Django
# ------------------------------------------------------------------
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from agendamentos.models import Professor, Sala, Agendamento

# ------------------------------------------------------------------
# Dependência: openpyxl
# ------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    sys.exit("❌  Instale openpyxl: pip install openpyxl")

# ------------------------------------------------------------------
# Constantes
# ------------------------------------------------------------------
ARQUIVO_XLSX = os.path.join(os.path.dirname(__file__), 'agendamentos_normalizados.xlsx')
ABA          = 'Agendamentos Normalizados'

# Colunas na planilha (linha 2 = cabeçalho, dados a partir da linha 3)
COL = {
    'dia_semana':    1,
    'data_inicio':   2,
    'data_fim':      3,
    'sala':          4,
    'horario_inicio':5,
    'horario_fim':   6,
    'qtd_alunos':    7,
    'professor':     8,
    'disciplina':    9,
    'temas':         10,
    'observacoes':   11,
}

DIAS_VALIDOS = {'SEG', 'TER', 'QUA', 'QUI', 'SEX'}


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def limpar(val):
    return str(val).strip() if val is not None else ''


def parse_time(val):
    """'08:00' ou objeto time → '08:00:00'"""
    if val is None:
        return None
    s = str(val).strip()
    for fmt in ['%H:%M:%S', '%H:%M']:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


def parse_date(val):
    """'12/02/2026' ou objeto date → date"""
    if val is None:
        return None
    if hasattr(val, 'year'):          # já é date/datetime
        return val.date() if hasattr(val, 'hour') else val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def val(ws, row, col_name):
    return ws.cell(row=row, column=COL[col_name]).value


# ------------------------------------------------------------------
# Importação principal
# ------------------------------------------------------------------

def importar(dry_run=False, reset=False):

    if not os.path.isfile(ARQUIVO_XLSX):
        sys.exit(f"❌  Arquivo não encontrado: {ARQUIVO_XLSX}\n"
                 "    Execute primeiro normalizar_planilha.py")

    wb = openpyxl.load_workbook(ARQUIVO_XLSX, data_only=True)
    if ABA not in wb.sheetnames:
        sys.exit(f"❌  Aba '{ABA}' não encontrada no arquivo.")
    ws = wb[ABA]

    if reset and not dry_run:
        print("🗑️  Apagando todos os agendamentos, professores e salas …")
        Agendamento.objects.all().delete()
        Professor.objects.all().delete()
        Sala.objects.all().delete()
        print("✨  Banco limpo.\n")

    salvos = 0
    ignorados = 0
    erros = 0

    # Linhas de dados começam na linha 3 (linha 1 = título, linha 2 = cabeçalho)
    for row in range(3, ws.max_row + 1):

        dia_semana = limpar(val(ws, row, 'dia_semana')).upper()
        if dia_semana not in DIAS_VALIDOS:
            continue                  # linha vazia ou inválida

        sala_nome    = limpar(val(ws, row, 'sala'))
        prof_nome    = limpar(val(ws, row, 'professor'))
        disciplina   = limpar(val(ws, row, 'disciplina'))
        temas_raw    = limpar(val(ws, row, 'temas'))
        obs_raw      = limpar(val(ws, row, 'observacoes'))
        qtd_alunos   = int(val(ws, row, 'qtd_alunos') or 0)
        hi           = parse_time(val(ws, row, 'horario_inicio'))
        hf           = parse_time(val(ws, row, 'horario_fim'))
        data_inicio  = parse_date(val(ws, row, 'data_inicio'))
        data_fim     = parse_date(val(ws, row, 'data_fim'))

        # Concatena temas + observacoes como texto único para o campo 'temas'
        temas_completo = '\n'.join(filter(None, [temas_raw, obs_raw]))

        if not sala_nome or not prof_nome or not hi or not hf:
            print(f"⚠️  Linha {row}: dados incompletos — ignorada.")
            ignorados += 1
            continue

        if dry_run:
            print(f"[DRY-RUN] {dia_semana} | {data_inicio or 'Semestre todo'} | "
                  f"{sala_nome[:18]} | {hi} às {hf} | "
                  f"{prof_nome[:25]} | {disciplina[:20]}")
            salvos += 1
            continue

        try:
            sala_obj, _  = Sala.objects.get_or_create(nome=sala_nome)
            prof_obj, _  = Professor.objects.get_or_create(nome=prof_nome)

            agendamento, created = Agendamento.objects.get_or_create(
                dia_semana     = dia_semana,
                sala           = sala_obj,
                horario_inicio = hi,
                data_inicio    = data_inicio,
                defaults={
                    'horario_fim': hf,
                    'qtd_alunos':  qtd_alunos,
                    'professor':   prof_obj,
                    'disciplina':  disciplina or None,
                    'data_fim':    data_fim,
                    'temas':       temas_completo or None,
                }
            )

            if created:
                dt_str = data_inicio.strftime('%d/%m') if data_inicio else 'Semestre todo'
                print(f"✅ Salvo  | {dia_semana} | {dt_str:<9} | "
                      f"{sala_nome[:18]:<18} | {prof_nome[:25]}")
                salvos += 1
            else:
                print(f"⏭️  Existe | {dia_semana} | "
                      f"{sala_nome[:18]:<18} | {prof_nome[:25]}")
                ignorados += 1

        except Exception as e:
            print(f"❌ Erro na linha {row}: {e}")
            erros += 1

    print(f"\n{'='*60}")
    if dry_run:
        print(f"[DRY-RUN] {salvos} registros seriam importados.")
    else:
        print(f"Resultado: {salvos} salvos | {ignorados} já existiam | {erros} erros")
    print('='*60)


# ------------------------------------------------------------------
# Entry-point
# ------------------------------------------------------------------
if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Importa agendamentos normalizados para o Django.')
    parser.add_argument('--dry-run', action='store_true',
                        help='Mostra o que seria importado sem salvar no banco.')
    parser.add_argument('--reset', action='store_true',
                        help='Apaga todos os dados existentes antes de importar.')
    args = parser.parse_args()

    importar(dry_run=args.dry_run, reset=args.reset)